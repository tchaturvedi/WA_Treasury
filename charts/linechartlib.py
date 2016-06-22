import json
from openpyxl import load_workbook

def createLineChart(charttype):
    with open("charts/chartconfig.json") as config:
        jsonData = json.load(config)

    chart = jsonData[charttype]
    chartTitle = chart['chart-title']
    dataTitle = chart['data-title']
    rowOffset = chart['row-offset']
    colors = chart['color']
    yFormat = chart['y-axis-format']

    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb[dataTitle]

    xaxis = chart['x-axis']
    items = chart['y-axis']
    data = []

    for item in items:
        key = None
        for k in item.keys():
            key = k
        dict = {
            "type": "line",
            "name": key,
            "cursor": "pointer",
            "showInLegend": True,
            "color": colors[0]
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=int(rowOffset)):
            d = {
                "label": row[xaxis].value,
                "y": row[item[key]].value
            }
            if d['label'] is not None:
                dataPoints.append(d)

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = chartTitle
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)
        chartInfo["valueFormat"] = yFormat
    return chartInfo