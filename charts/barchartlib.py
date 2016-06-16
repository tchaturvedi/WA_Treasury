import json
from openpyxl import load_workbook

def createBarChart(charttype):
    with open("charts/chartconfig.json") as config:
        jsonData = json.load(config)

    config.close()
    chart = jsonData[charttype]
    chartTitle = chart['chart-title']
    dataTitle = chart['data-title']
    rowOffset = chart['row-offset']

    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb[dataTitle]

    xaxis = chart['x-axis']
    items = chart['y-axis']
    colors = chart['color']
    colorIndex = 0
    data = []

    for item in items:
        key = None
        for k in item.keys():
            key = k
        dict = {
            "type": "stackedColumn",
            "legendText": key,
            "cursor": "pointer",
            "showInLegend": True,
            'legendMarkerColor': colors[colorIndex],
            "toolTipContent": key + " in year " + "{label}: {y}"
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=int(rowOffset)):
            if type(item[key]) is list:
                if row[item[key][1]] == key:
                    d = {
                        "label": row[xaxis].value,
                        "y": row[item[key][0]].value,
                        'color': colors[colorIndex]
                    }
            else:
                d = {
                    "label": row[xaxis].value,
                    "y": row[item[key]].value,
                    'color': colors[colorIndex]
                }
            if d['label'] is not None:
                dataPoints.append(d)
        colorIndex += 1

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = chartTitle
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)
    return chartInfo
