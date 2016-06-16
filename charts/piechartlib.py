import json
from openpyxl import load_workbook

class CreatePieChart():

    def __init__(self, year, chartType):
        self.year = year
        self.chartType = chartType

    def createPieChart(self):
        with open("charts/chartconfig.json") as config:
            data = json.load(config)

        chartconfig = {}
        for key, value in data.items():
            if key == self.chartType:
                chartconfig = value

        # The position index from the excel file
        xaxis = chartconfig['x-axis']
        # An array of dictionary with Name of the category and index in excel file
        yaxis = chartconfig['y-axis']

        charttitle = chartconfig['chart-title']
        datatitle = chartconfig['data-title']
        rowoffset = chartconfig['row-offset']
        colors = chartconfig['color']
        colorIndex = 0
        wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
        sheet = wb[datatitle]
        newData = []
        dic = {
            "type": "pie",
            "showInLegend": True,
            "toolTipContent": "{y} - #percent %",
            "yValueFormatString": "#0.#,,. Million",
            "legendText": "{indexLabel}",
        }
        dataPoints = {}
        for row in sheet.iter_rows(row_offset=int(rowoffset)):
            if row[xaxis].value == self.year:
                for cat in yaxis:
                    for key in cat.keys():
                        if row[cat[key]].value is not None and row[cat[key]].value > 0:
                            if key not in dataPoints.keys():
                                dataPoints[key] = {'y': row[cat[key]].value, 'indexLabel': key,
                                                   'legendMarkerColor': colors[colorIndex], 'color': colors[colorIndex]}
                            else:
                                dataPoints[key]['y'] += row[cat[key]].value
                    colorIndex += 1

        for item in dataPoints:
            newData.append(dataPoints[item])

        dic['dataPoints'] = newData
        pieChartInfo = {}
        pieChartInfo['title'] = charttitle + str(self.year)
        pieChartInfo['data'] = dic
        pieChartInfo['json'] = json.dumps(dic)
        return pieChartInfo