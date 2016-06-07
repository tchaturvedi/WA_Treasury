import json
from datetime import datetime

from flask import Flask, render_template, redirect, request, jsonify
from flask_flatpages import FlatPages
from openpyxl import load_workbook
from slugify import slugify
from util import buildChartInfo
from charts.piechartlib import CreatePieChart

######################
# Configurations
######################
import conf

######################
# Initialization
######################
app = Flask(__name__)
app.config.from_object(conf)
pages = FlatPages(app)

# generate the Table of content and page slugs from the title
# exclude the index page as its can be accessed by clicking on the page logo
TOC = [(page['title'], slugify(page['title'])) for page in pages if page['title'].lower() != 'index']

######################
# Routes
######################

@app.route('/')
def index():
    return redirect('index')


@app.route('/<path:path>')
def page(path):
    '''
        Entry point for each page based on the page slug
    '''
    page = pages.get_or_404(path)
    buildChartInfo(page.html)
    chartInfo = getChart1Info()
    return render_template('page.html', page=page, toc=TOC, year=datetime.now().year, chartInfo=chartInfo)


@app.route('/createPieChart', methods=['POST'])
def createPieChart():
    sentInData = request.get_json(force=True)
    chart = None
    for key in sentInData.keys():
        chart = key

    createPie = CreatePieChart(sentInData[chart]["label"], chart)
    return jsonify(createPie.createPieChart())

#####################
# Helper Functions
#####################

def getChart1Info():
    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb.get_sheet_by_name('Outstanding (Fig 2)')

    items = ["VP GO", "MVFT GO", "Triple Pledge", "GARVEEs", "TIFIA", "State COPs"]
    data = []

    for item in items:
        dict = {
            "type": "stackedColumn",
            "name": item,
            "cursor": "pointer",
            "showInLegend": True
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=3):
            if item == "VP GO":
                d = {
                    "label": row[0].value,
                    "y": row[1].value
                }
            elif item == "MVFT GO":
                d = {
                    "label": row[0].value,
                    "y": row[2].value
                }
            elif item == "Triple Pledge":
                d = {
                    "label": row[0].value,
                    "y": row[3].value
                }
            elif item == "GARVEEs":
                d = {
                    "label": row[0].value,
                    "y": row[4].value
                }
            elif item == "TIFIA":
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            else:
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            if d["label"] != None:
                dataPoints.append(d)

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = "Outstanding Bonds and COPs FY 2000-2016 ($ Billions)"
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)
    return chartInfo

if __name__ == '__main__':
    app.run(debug=conf.DEBUG)