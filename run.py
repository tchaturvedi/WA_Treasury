import json
import os

from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
from datetime import datetime
import markdown

BASEDIR = os.path.abspath(os.path.dirname(__file__))
CONTENT_DIR = os.path.abspath(os.path.join(os.sep, BASEDIR, 'content'))
CONTENT_EXTENSION = '.md'
encoding = 'utf-8'


pages = {}
app = Flask(__name__)


@app.route('/')
@app.route('/index.html')
def index():
    print(getChart1Info())
    return render_template(
        'index.html',
        title='Home Page',
        chart1Info=getChart1Info(),
        chart2Info=getChart2Info(),
        chart3Info=getChart3Info(),
        chart4Info=getChart4Info(),
        year=datetime.now().year,
    )


@app.route('/chart1.html')
def chart1():
    return render_template(
        'chart1.html',
        title='Chart 1',
        chart1Info=getChart1Info()
    )


@app.route('/chart2.html')
def chart2():
    return render_template(
        'chart2.html',
        title='Chart 2',
        chart2Info=getChart2Info()
    )

@app.route('/chart3.html')
def chart3():
    return render_template(
        'chart3.html',
        title='Chart 3',
        chart3Info=getChart3Info()
    )

@app.route('/chart4.html')
def chart4():
    return render_template(
        'chart4.html',
        title='Chart 4',
        chart4Info=getChart4Info()
    )




@app.route('/createPieChart', methods=["POST"])
def createPieChart():
    sentInData = request.get_json(force=True)
    chart = None
    for key in sentInData.keys():
        chart = key

    if chart == "chart1":
        return jsonify(getPieChartInfo(sentInData[chart]["label"], getChart1Info()))
    elif chart == "chart2":
        return jsonify(getPieChartInfo(sentInData[chart]["label"], getChart2Info()))
    elif chart == "chart3":
        return jsonify(getPieChartInfo(sentInData[chart]["label"], getChart3Info()))
    elif chart == "chart4":
        return jsonify(getPieChartInfo(sentInData[chart]["label"], getChart4Info()))


def getPieChartInfo(year, chartInfo):
    oldData = chartInfo["data"]
    newData = []
    dic = {
        "type": "pie",
        "showInLegend": True,
        "toolTipContent": "{y} - #percent %",
        "yValueFormatString": "#0.#,,. Million",
        "legendText": "{indexLabel}",
    }
    dataPoints = {}
    for item in oldData:
        for dataItem in item["dataPoints"]:
            if dataItem["label"] == year:
                if item["name"] not in dataPoints.keys():
                    dataPoints[item["name"]] = {"y": dataItem["y"], "indexLabel": item["name"]}
                else:
                    dataPoints[item["name"]]["y"] += dataItem["y"]

    for item in dataPoints:
        newData.append(dataPoints[item])

    dic["dataPoints"] = newData
    pieChartInfo = {}
    pieChartInfo["title"] = "Detailed Data " + str(year)
    pieChartInfo["data"] = dic
    pieChartInfo["json"] = json.dumps(dic)
    return pieChartInfo


def getChart1Info():
    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb.get_sheet_by_name('Outstanding (Fig 2)')

    items = ["VP GO", "MVFT GO", "Triple Pledge", "GARVEEs", "TIFIA", "State COPs"]
    data = []

    for item in items:
        dict = {
            "type": "stackedColumn",
            "name": item,
            "cursor": "pointer",
            "showInLegend": True
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=3):
            if item == "VP GO":
                d = {
                    "label": row[0].value,
                    "y": row[1].value
                }
            elif item == "MVFT GO":
                d = {
                    "label": row[0].value,
                    "y": row[2].value
                }
            elif item == "Triple Pledge":
                d = {
                    "label": row[0].value,
                    "y": row[3].value
                }
            elif item == "GARVEEs":
                d = {
                    "label": row[0].value,
                    "y": row[4].value
                }
            elif item == "TIFIA":
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            else:
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            if d["label"] != None:
                dataPoints.append(d)

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = "Outstanding Bonds and COPs FY 2000-2016 ($ Billions)"
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)

        page = getContent()['Debt_Portfolio_Overview']
        chartInfo['content'] = page

    return chartInfo


def getChart2Info():
    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb.get_sheet_by_name('New Money Issuance (Fig 3)')

    items = ["VP GO", "MVFT GO", "Triple Pledge", "GARVEEs", "TIFIA", "State COPs"]
    data = []

    for item in items:
        dict = {
            "type": "stackedColumn",
            "name": item,
            "cursor": "pointer",
            "showInLegend": True
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=2):
            if item == "VP GO":
                d = {
                    "label": row[0].value,
                    "y": row[1].value
                }
            elif item == "MVFT GO":
                d = {
                    "label": row[0].value,
                    "y": row[2].value
                }
            elif item == "Triple Pledge":
                d = {
                    "label": row[0].value,
                    "y": row[3].value
                }
            elif item == "GARVEEs":
                d = {
                    "label": row[0].value,
                    "y": row[4].value
                }
            elif item == "TIFIA":
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            else:
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            if d["label"] != None:
                dataPoints.append(d)

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = "Bond and COP Issuance FY 2000-2016 ($ Millions)"
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)

    return chartInfo

def getChart3Info():
    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb.get_sheet_by_name('Debt Service (Fig 4,5)')

    items = ["VP GO", "MVFT GO", "Triple Pledge", "GARVEEs", "TIFIA", "State COPs"]
    data = []

    for item in items:
        dict = {
            "type": "stackedColumn",
            "name": item,
            "cursor": "pointer",
            "showInLegend": True
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=2):
            if item == "VP GO":
                d = {
                    "label": row[0].value,
                    "y": row[1].value
                }
            elif item == "MVFT GO":
                d = {
                    "label": row[0].value,
                    "y": row[2].value
                }
            elif item == "Triple Pledge":
                d = {
                    "label": row[0].value,
                    "y": row[3].value
                }
            elif item == "GARVEEs":
                d = {
                    "label": row[0].value,
                    "y": row[4].value
                }
            elif item == "TIFIA":
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            else:
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            if d["label"] != None:
                dataPoints.append(d)

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = "Debt Service Paid and Due 2000 - 2016 ($ Millions)"
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)

    return chartInfo

def getChart4Info():
    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb.get_sheet_by_name('Debt Serv % of Gen Fund (Fig10)')

    items = ["Debt Service as % of GF-S Revenues"]
    data = []

    for item in items:
        dict = {
            "type": "line",
            "name": item,
            "cursor": "pointer",
            "showInLegend": True
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=3):
            if item == "Debt Service as % of GF-S Revenues":
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            else:
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            if d["label"] != None:
                dataPoints.append(d)

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = "Debt Serv % of Gen Fund "
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)

    return chartInfo


def getContent():
    pages = {}
    for root, dirs, files in os.walk(CONTENT_DIR):
        for name in files:
            if not name.endswith(CONTENT_EXTENSION):
                continue

            filename = os.path.join(root, name)
            with open(filename) as file:
                content = file.read()

            pages[name[:-len(CONTENT_EXTENSION)]] = markdown.markdown(content)

    return pages

def _parse(content):
    lines = iter(content.split('\n'))
    # handle metadata in the file
    content = '\n'.join(lines)
    return content

if __name__ == '__main__':
    app.debug = True
    app.run()